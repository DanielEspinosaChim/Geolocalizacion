import io
import json
import re
from datetime import datetime
from difflib import SequenceMatcher
from fastapi import APIRouter, HTTPException, Body, UploadFile, File
from fastapi.responses import StreamingResponse

from backend.core.firebase import fdb

router = APIRouter()

MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

CATEGORY_ORDER = ['FRUTAS', 'VEGETALES', 'ABARROTES', 'CARNES', 'LECHES', 'HIGIENE', 'FARMACÉUTICOS']

_SEED_2026 = [
    {"id": "aguacate",        "name": "AGUACATE",               "category": "FRUTAS",         "unit": "KILO",    "sort_order": 1,
     "prices": {"jan": 39.95, "feb": 49.95, "mar": 46.95, "apr": 46.95, "may": 54.95, "jun": 124.00, "jul": 79.00}},
    {"id": "limon",           "name": "LIMON",                  "category": "FRUTAS",         "unit": "KILO",    "sort_order": 2,
     "prices": {"jan": 29.95, "feb": 29.95, "mar": 56.95, "apr": 76.95, "may": 49.95, "jun": 39.95,  "jul": 39.95}},
    {"id": "cebolla",         "name": "CEBOLLA",                "category": "VEGETALES",      "unit": "KILO",    "sort_order": 3,
     "prices": {"jan": 46.95, "feb": 46.95, "mar": 41.93, "apr": 29.95, "may": 29.95, "jun": 24.50,  "jul": 32.50}},
    {"id": "jitomate",        "name": "JITOMATE",               "category": "VEGETALES",      "unit": "KILO",    "sort_order": 4,
     "prices": {"jan": 32.95, "feb": 32.95, "mar": 59.00, "apr": 56.95, "may": 74.95, "jun": 52.50,  "jul": 32.95}},
    {"id": "aceite_vegetal",  "name": "ACEITE VEGETAL",         "category": "ABARROTES",      "unit": "LITRO",   "sort_order": 5,
     "prices": {"jan": 46.95, "feb": 46.95, "mar": 46.95, "apr": 57.90, "may": 42.95, "jun": 51.95,  "jul": 51.95}},
    {"id": "arroz",           "name": "ARROZ",                  "category": "ABARROTES",      "unit": "900 GR.", "sort_order": 6,
     "prices": {"jan": 50.50, "feb": 50.50, "mar": 50.50, "apr": 50.50, "may": 51.95, "jun": 51.95,  "jul": 52.95}},
    {"id": "azucar",          "name": "AZUCAR",                 "category": "ABARROTES",      "unit": "KILO",    "sort_order": 7,
     "prices": {"jan": 35.00, "feb": 35.00, "mar": 35.00, "apr": 35.00, "may": 35.00, "jun": 35.00,  "jul": 35.00}},
    {"id": "frijol",          "name": "FRIJOL",                 "category": "ABARROTES",      "unit": "900 GR",  "sort_order": 8,
     "prices": {"jan": 36.50, "feb": 36.50, "mar": 34.00, "apr": 34.00, "may": 34.00, "jun": 34.00,  "jul": 23.95}},
    {"id": "tortillas",       "name": "TORTILLAS",              "category": "ABARROTES",      "unit": "KILO",    "sort_order": 9,
     "prices": {"jan": 32.00, "feb": 32.00, "mar": 32.00, "apr": 32.00, "may": 32.00, "jun": 32.00,  "jul": 32.00}},
    {"id": "res_bistec",      "name": "RES BISTEC ESPECIAL",    "category": "CARNES",         "unit": "KILO",    "sort_order": 10,
     "prices": {"jan": 245.00, "feb": 278.00, "mar": 279.00, "apr": 254.00, "may": 195.00, "jun": 245.00, "jul": 275.00}},
    {"id": "cerdo_lomo",      "name": "CERDO LOMO",             "category": "CARNES",         "unit": "KILO",    "sort_order": 11,
     "prices": {"jan": 118.00, "feb": 124.00, "mar": 105.00, "apr": 109.00, "may": 102.00, "jun": 114.00, "jul": 108.00}},
    {"id": "huevo",           "name": "HUEVO (REJILLA 30 PZAS)","category": "CARNES",         "unit": "REJILLA", "sort_order": 12,
     "prices": {"jan": 96.50, "feb": 96.50, "mar": 96.50, "apr": 96.50, "may": 89.95, "jun": 89.95,  "jul": 89.95}},
    {"id": "pollo_pechuga",   "name": "POLLO PECHUGA",          "category": "CARNES",         "unit": "KILO",    "sort_order": 13,
     "prices": {"jan": 155.00, "feb": 155.00, "mar": 189.00, "apr": 155.00, "may": 185.00, "jun": 199.00, "jul": 165.00}},
    {"id": "leche_entera",    "name": "LECHE ENTERA",           "category": "LECHES",         "unit": "1LT",     "sort_order": 14,
     "prices": {"jan": 26.25, "feb": 26.25, "mar": 26.25, "apr": 38.95, "may": 38.95, "jun": 35.50,  "jul": 38.95}},
    {"id": "jabon_tocador",   "name": "JABON DE TOCADOR",       "category": "HIGIENE",        "unit": "1PZA",    "sort_order": 15,
     "prices": {"jan": 11.00, "feb": 11.00, "mar": 11.64, "apr": 11.64, "may": 9.91,  "jun": 15.00,  "jul": 15.00}},
    {"id": "papel_higienico", "name": "PAPEL HIGIENICO",        "category": "HIGIENE",        "unit": "4 PZA",   "sort_order": 16,
     "prices": {"jan": 119.83, "feb": 119.83, "mar": 119.83, "apr": 119.83, "may": 115.52, "jun": 111.21, "jul": 111.21}},
    {"id": "detergente",      "name": "DETERGENTE",             "category": "HIGIENE",        "unit": "750 GR",  "sort_order": 17,
     "prices": {"jan": 28.00, "feb": 28.00, "mar": 32.76, "apr": 30.50, "may": 30.50, "jun": 33.00,  "jul": 33.00}},
    {"id": "algodon",         "name": "ALGODON",                "category": "FARMACÉUTICOS",  "unit": "500 GR",  "sort_order": 18,
     "prices": {"jan": 104.99, "feb": 104.99, "mar": 104.99, "apr": 104.99, "may": 104.99, "jun": 104.99, "jul": 104.99}},
    {"id": "agua_oxigenada",  "name": "AGUA OXIGENADA",         "category": "FARMACÉUTICOS",  "unit": "LITRO",   "sort_order": 19,
     "prices": {"jan": 49.00, "feb": 49.00, "mar": 49.00, "apr": 49.00, "may": 49.00, "jun": 49.00,  "jul": 49.00}},
]


def _col(year: str):
    return fdb.collection("canasta_basica").document(year).collection("products")


def _get_products(year: str) -> list:
    docs = list(_col(year).where("active", "==", True).stream())
    products = []
    for doc in docs:
        d = doc.to_dict()
        d["id"] = doc.id
        products.append(d)
    products.sort(key=lambda p: (
        CATEGORY_ORDER.index(p.get("category", "")) if p.get("category") in CATEGORY_ORDER else 99,
        p.get("sort_order", 99)
    ))
    return products


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get(
    "/api/canasta/{year}",
    tags=["Canasta Básica"],
    summary="Listar productos del año",
    description=(
        "Retorna todos los productos activos del año indicado ordenados por categoría "
        "y sort_order. Cada producto incluye sus 12 meses de precios (null = sin capturar)."
    ),
)
def get_canasta(year: str):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    try:
        return _get_products(year)
    except Exception as e:
        raise HTTPException(500, str(e))


@router.put(
    "/api/canasta/{year}/{product_id}",
    tags=["Canasta Básica"],
    summary="Actualizar precio de un mes",
    description=(
        "Actualiza o borra el precio de un producto para el mes indicado. "
        "Si ya existía un valor lo sobreescribe. "
        "Enviar `price: null` borra el valor del mes. "
        "Campos opcionales: `tienda` (string) y `fecha_compra` (YYYY-MM-DD). "
        "Meses válidos: jan feb mar apr may jun jul aug sep oct nov dec."
    ),
)
def update_precio(year: str, product_id: str, body: dict = Body(
    ...,
    example={"month": "aug", "price": 85.00, "tienda": "WALMART", "fecha_compra": "2026-08-05"},
)):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    month = body.get("month")
    price = body.get("price")
    if month not in MONTHS:
        raise HTTPException(400, f"Mes inválido: {month}")
    try:
        price_val = float(price) if price not in (None, "", "-") else None
        update_data = {
            f"prices.{month}": price_val,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        # Tienda y fecha_compra son opcionales — solo se actualizan si vienen en el body
        if "tienda" in body:
            tienda = body["tienda"]
            update_data[f"tiendas.{month}"] = tienda.strip().upper() if tienda else None
        if "fecha_compra" in body:
            fecha = body["fecha_compra"]
            update_data[f"fechas_compra.{month}"] = fecha if fecha else None
        _col(year).document(product_id).update(update_data)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post(
    "/api/canasta/{year}/product",
    tags=["Canasta Básica"],
    summary="Agregar nuevo producto al catálogo",
    description=(
        "Crea un producto nuevo en el año indicado. El ID se genera automáticamente "
        "a partir del nombre (minúsculas, guiones bajos). "
        "El producto se crea con todos los meses en null. "
        "Categorías válidas: FRUTAS · VEGETALES · ABARROTES · CARNES · LECHES · HIGIENE · FARMACÉUTICOS."
    ),
)
def add_product(year: str, body: dict = Body(
    ...,
    example={"name": "PAPAYA", "category": "FRUTAS", "unit": "KILO"},
)):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    name = (body.get("name") or "").strip().upper()
    category = (body.get("category") or "").strip().upper()
    unit = (body.get("unit") or "").strip().upper()
    if not name or not category:
        raise HTTPException(400, "name y category son requeridos")
    try:
        existing = list(_col(year).stream())
        max_order = max((d.to_dict().get("sort_order", 0) for d in existing), default=0)
        product_id = name.lower().replace(" ", "_").replace("(", "").replace(")", "")
        _col(year).document(product_id).set({
            "name": name,
            "category": category,
            "unit": unit,
            "sort_order": max_order + 1,
            "active": True,
            "prices": {m: None for m in MONTHS},
            "tiendas": {},
            "fechas_compra": {},
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        return {"ok": True, "id": product_id}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.delete(
    "/api/canasta/{year}/{product_id}",
    tags=["Canasta Básica"],
    summary="Desactivar producto (soft delete)",
    description=(
        "Marca el producto como inactivo (active=false). "
        "No elimina el documento de Firestore — los datos históricos se conservan. "
        "El producto deja de aparecer en la tabla."
    ),
)
def delete_product(year: str, product_id: str):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    try:
        _col(year).document(product_id).update({"active": False})
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get(
    "/api/canasta/{year}/summary",
    tags=["Canasta Básica"],
    summary="Resumen mensual: totales y variación %",
    description=(
        "Calcula el costo total de la canasta por mes y la variación porcentual "
        "respecto al mes anterior. Los meses sin datos retornan null en todos los campos. "
        "Útil para la gráfica de barras y las filas de totales de la tabla."
    ),
)
def get_summary(year: str):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    try:
        products = _get_products(year)
        summary = []
        prev_total = None
        for month in MONTHS:
            prices = [p["prices"].get(month) for p in products if p.get("prices", {}).get(month) is not None]
            if not prices:
                summary.append({"month": month, "total": None, "diff": None, "pct": None})
                prev_total = None
                continue
            total = round(sum(prices), 2)
            diff = round(total - prev_total, 2) if prev_total is not None else None
            pct = round((diff / prev_total) * 100, 1) if prev_total is not None and prev_total != 0 else None
            summary.append({"month": month, "total": total, "diff": diff, "pct": pct})
            prev_total = total
        return summary
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post(
    "/api/canasta/{year}/seed",
    tags=["Canasta Básica"],
    summary="Cargar datos históricos 2026 (solo primera vez)",
    description=(
        "Inserta los 19 productos base con precios históricos ENE–JUL 2026. "
        "Solo disponible para year='2026'. Operación idempotente: omite los productos "
        "que ya existen. Usar únicamente al inicializar el módulo por primera vez."
    ),
)
def seed_canasta(year: str):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    if year != "2026":
        raise HTTPException(400, "Seed solo disponible para 2026")
    try:
        col = _col(year)
        existing = [d.id for d in col.stream()]
        inserted = 0
        for p in _SEED_2026:
            if p["id"] in existing:
                continue
            doc = {
                "name": p["name"],
                "category": p["category"],
                "unit": p["unit"],
                "sort_order": p["sort_order"],
                "active": True,
                "prices": {**{m: None for m in MONTHS}, **p["prices"]},
                "tiendas": {},
                "fechas_compra": {},
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            col.document(p["id"]).set(doc)
            inserted += 1
        return {"ok": True, "insertados": inserted, "ya_existian": len(existing)}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Exportación Excel ─────────────────────────────────────────────────────────

_MONTH_NAMES = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
                'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']

@router.get(
    "/api/canasta/{year}/export/excel",
    tags=["Canasta Básica"],
    summary="Descargar Excel (.xlsx)",
    description=(
        "Genera y descarga un archivo Excel con el formato original de CANACO SERVYTUR. "
        "Incluye: encabezado azul oscuro, filas coloreadas por categoría, "
        "fila TOTAL, fila DIFERENCIA (rojo/verde) y fila % CAMBIO. "
        "Solo se incluyen los meses que tienen al menos un precio capturado. "
        "Primeras 3 columnas + encabezado congelados (D2)."
    ),
)
def export_excel(year: str, compare: str = None):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(503, "Instala openpyxl: pip install openpyxl")

    try:
        products = _get_products(year)
        # Meses con al menos un precio
        active_months = [m for m in MONTHS if any(p.get("prices", {}).get(m) is not None for p in products)]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Canasta {year}"

        # ── Estilos ──────────────────────────────────────────────────────────
        hdr_fill  = PatternFill("solid", fgColor="1F3864")
        hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        cat_font  = Font(bold=True, size=10)
        num_fmt   = '#,##0.00'
        thin      = Side(style="thin", color="D0D7E8")
        border    = Border(bottom=thin)
        center    = Alignment(horizontal="center", vertical="center")
        right_al  = Alignment(horizontal="right")

        CAT_FILLS = {
            "FRUTAS":         "EBF5FB",
            "VEGETALES":      "E9F7EF",
            "ABARROTES":      "FEF9E7",
            "CARNES":         "FDEDEC",
            "LECHES":         "F4ECF7",
            "HIGIENE":        "EAF2FF",
            "FARMACÉUTICOS":  "FDF2E9",
        }

        # ── Encabezado ───────────────────────────────────────────────────────
        headers = ["CATEGORIA", "SUMINISTRO", "PRESENTACION"] + \
                  [_MONTH_NAMES[MONTHS.index(m)] for m in active_months]
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_i, value=h)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = border

        # Anchos de columna
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 12
        for ci in range(4, 4 + len(active_months)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 11

        # ── Filas de productos ───────────────────────────────────────────────
        last_cat = None
        row      = 2

        for p in products:
            cat = p.get("category", "")
            fill_hex = CAT_FILLS.get(cat, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_hex)
            is_new   = cat != last_cat
            last_cat = cat

            ws.cell(row=row, column=1, value=cat).font = cat_font if is_new else Font(color="888888", size=10)
            ws.cell(row=row, column=1).fill = row_fill
            ws.cell(row=row, column=2, value=p["name"]).fill = row_fill
            ws.cell(row=row, column=3, value=p.get("unit","")).fill = row_fill
            ws.cell(row=row, column=3).alignment = center

            for ci, m in enumerate(active_months, 4):
                val  = p.get("prices", {}).get(m)
                cell = ws.cell(row=row, column=ci, value=val if val is not None else "")
                cell.fill      = row_fill
                cell.alignment = right_al
                if val is not None:
                    cell.number_format = num_fmt

            row += 1

        # ── Fila TOTAL ───────────────────────────────────────────────────────
        totales = []
        ws.cell(row=row, column=1, value="").fill = PatternFill("solid", fgColor="1F3864")
        ws.cell(row=row, column=2, value="").fill = PatternFill("solid", fgColor="1F3864")
        ws.cell(row=row, column=3, value="TOTAL").font = Font(color="FFFFFF", bold=True, size=10)
        ws.cell(row=row, column=3).fill      = PatternFill("solid", fgColor="1F3864")
        ws.cell(row=row, column=3).alignment = center

        for ci, m in enumerate(active_months, 4):
            vals  = [p["prices"][m] for p in products if p.get("prices", {}).get(m) is not None]
            total = round(sum(vals), 2) if vals else None
            totales.append(total)
            cell  = ws.cell(row=row, column=ci, value=total if total is not None else "")
            cell.fill          = PatternFill("solid", fgColor="1F3864")
            cell.font          = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment     = right_al
            if total is not None:
                cell.number_format = num_fmt
        row += 1

        # ── Fila DIFERENCIA ──────────────────────────────────────────────────
        ws.cell(row=row, column=3, value="DIFERENCIA").font = Font(bold=True, size=10)
        ws.cell(row=row, column=3).alignment = center
        for ci, (total, prev) in enumerate(zip(totales, [None] + totales[:-1]), 4):
            if total is None or prev is None:
                continue
            diff  = round(total - prev, 2)
            cell  = ws.cell(row=row, column=ci, value=diff)
            cell.number_format = '+#,##0.00;-#,##0.00'
            cell.alignment     = right_al
            cell.font          = Font(color="C00000" if diff > 0 else "00B050", bold=True, size=10)
        row += 1

        # ── Fila % ───────────────────────────────────────────────────────────
        ws.cell(row=row, column=3, value="% CAMBIO").font = Font(bold=True, size=10)
        ws.cell(row=row, column=3).alignment = center
        for ci, (total, prev) in enumerate(zip(totales, [None] + totales[:-1]), 4):
            if total is None or prev is None or prev == 0:
                continue
            pct   = round((total - prev) / prev * 100, 1)
            cell  = ws.cell(row=row, column=ci, value=pct / 100)
            cell.number_format = '+0.0%;-0.0%'
            cell.alignment     = right_al
            cell.font          = Font(color="C00000" if pct > 0 else "00B050", bold=True, size=10)

        # ── Congelar encabezado ──────────────────────────────────────────────
        ws.freeze_panes = "D2"

        # ── Hoja "Tienda y Fecha" — misma estructura que precios ────────────────
        ws2 = wb.create_sheet("Tienda y Fecha")
        hdr_border2 = Border(bottom=Side(style="medium", color="1E3A6E"))
        for col_i, h in enumerate(headers, 1):
            cell2 = ws2.cell(row=1, column=col_i, value=h)
            cell2.fill      = hdr_fill
            cell2.font      = hdr_font
            cell2.alignment = center
            cell2.border    = hdr_border2

        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 28
        ws2.column_dimensions["C"].width = 12
        for ci2 in range(4, 4 + len(active_months)):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(ci2)].width = 20

        last_cat2 = None
        row2 = 2
        for p in products:
            cat = p.get("category", "")
            fill_hex2 = CAT_FILLS.get(cat, "FFFFFF")
            row_fill2 = PatternFill("solid", fgColor=fill_hex2)
            is_new2   = cat != last_cat2
            last_cat2 = cat

            side_border = Border(bottom=thin, right=Side(style="thin", color="D1D9F0"))
            c2a = ws2.cell(row=row2, column=1, value=cat if is_new2 else "")
            c2a.font      = cat_font if is_new2 else Font(color="AAAAAA", size=9)
            c2a.fill      = row_fill2
            c2a.border    = side_border
            c2b = ws2.cell(row=row2, column=2, value=p["name"])
            c2b.fill      = row_fill2
            c2b.font      = Font(size=10, bold=True, color="1E293B")
            c2b.border    = side_border
            c2c = ws2.cell(row=row2, column=3, value=p.get("unit", ""))
            c2c.fill      = row_fill2
            c2c.alignment = center
            c2c.font      = Font(size=9, color="64748B")
            c2c.border    = Border(bottom=thin, right=Side(style="medium", color="B0B8D0"))

            for ci2, m in enumerate(active_months, 4):
                tienda_v = ((p.get("tiendas") or {}).get(m) or "").strip()
                fecha_v  = (p.get("fechas_compra") or {}).get(m) or ""
                try:
                    dia = str(int(fecha_v.split("-")[2])) if fecha_v else ""
                except Exception:
                    dia = ""
                if tienda_v and dia:
                    val2 = f"{tienda_v}  ·  Día {dia}"
                elif tienda_v:
                    val2 = tienda_v
                elif dia:
                    val2 = f"Día {dia}"
                else:
                    val2 = "—"
                c2 = ws2.cell(row=row2, column=ci2, value=val2)
                c2.fill      = row_fill2
                c2.alignment = center
                c2.border    = Border(bottom=thin, right=Side(style="thin", color="E0E7FF"))
                c2.font      = Font(
                    size=10,
                    color="2563EB" if (tienda_v or dia) else "C8D3E8",
                    bold=bool(tienda_v),
                )

            row2 += 1

        ws2.freeze_panes = "D2"
        ws2.row_dimensions[1].height = 22

        # ── Hoja "Variación %" ───────────────────────────────────────────────
        ws3 = wb.create_sheet("Variación %")
        for col_i, h in enumerate(headers, 1):
            c = ws3.cell(row=1, column=col_i, value=h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = center
            c.border = Border(bottom=Side(style="medium", color="1E3A6E"))

        ws3.column_dimensions["A"].width = 16
        ws3.column_dimensions["B"].width = 26
        ws3.column_dimensions["C"].width = 12
        for ci in range(4, 4 + len(active_months)):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 13

        row3 = 2
        last_cat3 = None
        tot3_rows = []
        for p in products:
            cat = p.get("category", "")
            fill_hex3 = CAT_FILLS.get(cat, "FFFFFF")
            rf3 = PatternFill("solid", fgColor=fill_hex3)
            is_new3 = cat != last_cat3; last_cat3 = cat
            ws3.cell(row=row3, column=1, value=cat if is_new3 else "").fill = rf3
            ws3.cell(row=row3, column=1).font = cat_font if is_new3 else Font(color="AAAAAA", size=9)
            ws3.cell(row=row3, column=2, value=p["name"]).fill = rf3
            ws3.cell(row=row3, column=2).font = Font(size=10, bold=True, color="1E293B")
            ws3.cell(row=row3, column=3, value=p.get("unit","")).fill = rf3
            ws3.cell(row=row3, column=3).alignment = center
            ws3.cell(row=row3, column=3).font = Font(size=9, color="64748B")

            row_pcts = []
            for i, (ci, m) in enumerate(zip(range(4, 4 + len(active_months)), active_months)):
                val  = p.get("prices", {}).get(m)
                prev = p.get("prices", {}).get(active_months[i-1]) if i > 0 else None
                c3   = ws3.cell(row=row3, column=ci)
                c3.fill = rf3
                c3.alignment = right_al
                c3.border = Border(bottom=thin, right=Side(style="thin", color="E0E7FF"))
                if i > 0 and val is not None and prev is not None and prev != 0:
                    pct = (val - prev) / prev
                    c3.value = pct
                    c3.number_format = '+0.0%;-0.0%'
                    c3.font = Font(color="C00000" if pct > 0 else "00B050", bold=True, size=10)
                    row_pcts.append(pct)
                else:
                    c3.value = "—" if i == 0 else ""
                    c3.font = Font(color="CCCCCC", size=9)
            tot3_rows.append(row_pcts)
            row3 += 1

        # Fila total variación
        ws3.cell(row=row3, column=3, value="% TOTAL CANASTA").font = Font(bold=True, size=10, color="FFFFFF")
        ws3.cell(row=row3, column=3).fill = PatternFill("solid", fgColor="1F3864")
        ws3.cell(row=row3, column=3).alignment = center
        for i, (ci, m) in enumerate(zip(range(4, 4 + len(active_months)), active_months)):
            t  = totales[i] if i < len(totales) else None
            tp = totales[i-1] if i > 0 and (i-1) < len(totales) else None
            c3 = ws3.cell(row=row3, column=ci)
            c3.fill = PatternFill("solid", fgColor="1F3864")
            if t and tp and tp != 0:
                pct = (t - tp) / tp
                c3.value = pct
                c3.number_format = '+0.0%;-0.0%'
                c3.font = Font(color="FF9999" if pct > 0 else "99FFB3", bold=True, size=10)
                c3.alignment = right_al
        ws3.freeze_panes = "D2"

        # ── Hoja "Trimestres" ────────────────────────────────────────────────
        QUARTERS_DEF = [
            ("Q1  ENE–MAR", ["jan","feb","mar"]),
            ("Q2  ABR–JUN", ["apr","may","jun"]),
            ("Q3  JUL–SEP", ["jul","aug","sep"]),
            ("Q4  OCT–DIC", ["oct","nov","dec"]),
        ]
        active_q = [(lbl, [m for m in months if m in active_months])
                    for lbl, months in QUARTERS_DEF
                    if any(m in active_months for m in months)]
        if len(active_q) > 1:
            all_qm = [m for _, ms in active_q for m in ms]
            active_q.append(("PROMEDIO ANUAL", all_qm))

        ws4 = wb.create_sheet("Trimestres")
        hdr4 = ["CATEGORIA","SUMINISTRO","PRESENTACION"] + [lbl for lbl, _ in active_q]
        for col_i, h in enumerate(hdr4, 1):
            c = ws4.cell(row=1, column=col_i, value=h)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = center
            c.border = Border(bottom=Side(style="medium", color="1E3A6E"))

        ws4.column_dimensions["A"].width = 16
        ws4.column_dimensions["B"].width = 26
        ws4.column_dimensions["C"].width = 12
        for ci in range(4, 4 + len(active_q)):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18

        row4 = 2
        last_cat4 = None
        qtotals = [0.0] * len(active_q)
        qcounts = [0]   * len(active_q)
        for p in products:
            cat = p.get("category","")
            rf4 = PatternFill("solid", fgColor=CAT_FILLS.get(cat,"FFFFFF"))
            is_new4 = cat != last_cat4; last_cat4 = cat
            ws4.cell(row=row4, column=1, value=cat if is_new4 else "").fill = rf4
            ws4.cell(row=row4, column=1).font = cat_font if is_new4 else Font(color="AAAAAA", size=9)
            ws4.cell(row=row4, column=2, value=p["name"]).fill = rf4
            ws4.cell(row=row4, column=2).font = Font(size=10, bold=True, color="1E293B")
            ws4.cell(row=row4, column=3, value=p.get("unit","")).fill = rf4
            ws4.cell(row=row4, column=3).alignment = center

            for qi, (_, qmonths) in enumerate(active_q):
                vals = [p.get("prices",{}).get(m) for m in qmonths if p.get("prices",{}).get(m) is not None]
                c4 = ws4.cell(row=row4, column=4+qi)
                c4.fill = rf4
                c4.alignment = right_al
                c4.border = Border(bottom=thin, right=Side(style="thin", color="E0E7FF"))
                if vals:
                    avg = round(sum(vals)/len(vals), 2)
                    c4.value = avg
                    c4.number_format = num_fmt
                    c4.font = Font(size=10, color="1E293B")
                    qtotals[qi] += avg
                    qcounts[qi] += 1
                else:
                    c4.value = ""; c4.font = Font(color="CCCCCC", size=9)
            row4 += 1

        # Fila total trimestres
        ws4.cell(row=row4, column=3, value="TOTAL CANASTA").font = Font(color="FFFFFF", bold=True, size=10)
        ws4.cell(row=row4, column=3).fill = PatternFill("solid", fgColor="1F3864")
        ws4.cell(row=row4, column=3).alignment = center
        for qi in range(len(active_q)):
            c4 = ws4.cell(row=row4, column=4+qi, value=round(qtotals[qi],2) if qcounts[qi] else "")
            c4.fill = PatternFill("solid", fgColor="1F3864")
            c4.font = Font(color="FFFFFF", bold=True, size=10)
            c4.alignment = right_al
            if qcounts[qi]: c4.number_format = num_fmt
        ws4.freeze_panes = "D2"

        # ── Hoja "vs {compare}" — solo si se pidió comparación ───────────────
        if compare and compare != year:
            products_b = _get_products(compare)
            prices_b   = {p["id"]: p.get("prices", {}) for p in products_b}

            ws5 = wb.create_sheet(f"{year} vs {compare}")
            for col_i, h in enumerate(headers, 1):
                c = ws5.cell(row=1, column=col_i, value=h)
                c.fill = hdr_fill; c.font = hdr_font; c.alignment = center
                c.border = Border(bottom=Side(style="medium", color="1E3A6E"))
            # Sub-encabezado con años
            ws5.cell(row=2, column=3, value=f"← {year}  ·  {compare} →").font = Font(size=9, color="64748B", italic=True)
            ws5.cell(row=2, column=3).alignment = center

            ws5.column_dimensions["A"].width = 16
            ws5.column_dimensions["B"].width = 26
            ws5.column_dimensions["C"].width = 12
            for ci in range(4, 4 + len(active_months)):
                ws5.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20

            row5 = 3
            last_cat5 = None
            for p in products:
                cat = p.get("category", "")
                rf5 = PatternFill("solid", fgColor=CAT_FILLS.get(cat, "FFFFFF"))
                is_new5 = cat != last_cat5; last_cat5 = cat
                ws5.cell(row=row5, column=1, value=cat if is_new5 else "").fill = rf5
                ws5.cell(row=row5, column=1).font = cat_font if is_new5 else Font(color="AAAAAA", size=9)
                ws5.cell(row=row5, column=2, value=p["name"]).fill = rf5
                ws5.cell(row=row5, column=2).font = Font(size=10, bold=True, color="1E293B")
                ws5.cell(row=row5, column=3, value=p.get("unit","")).fill = rf5
                ws5.cell(row=row5, column=3).alignment = center

                pb = prices_b.get(p["id"], {})
                for ci, m in enumerate(active_months, 4):
                    va = p.get("prices", {}).get(m)
                    vb = pb.get(m)
                    c5 = ws5.cell(row=row5, column=ci)
                    c5.fill   = rf5
                    c5.border = Border(bottom=thin, right=Side(style="thin", color="E0E7FF"))
                    if va is not None and vb and vb != 0:
                        pct = (va - vb) / vb
                        label = f"${va:.2f}  /  ${vb:.2f}  ({'+' if pct>0 else ''}{pct*100:.1f}%)"
                        c5.value     = label
                        c5.font      = Font(size=9,
                                            color="C00000" if pct > 0 else ("00B050" if pct < 0 else "475569"),
                                            bold=True)
                        c5.alignment = center
                    elif va is not None:
                        c5.value = f"${va:.2f}  /  s/d"
                        c5.font  = Font(size=9, color="94A3B8")
                        c5.alignment = center
                    else:
                        c5.value = ""; c5.font = Font(color="CCCCCC", size=9)
                row5 += 1

            # Fila totales comparación
            ws5.cell(row=row5, column=3, value="TOTAL CANASTA").font = Font(color="FFFFFF", bold=True, size=10)
            ws5.cell(row=row5, column=3).fill = PatternFill("solid", fgColor="1F3864")
            ws5.cell(row=row5, column=3).alignment = center
            for ci, (m, ta) in enumerate(zip(active_months, totales), 4):
                vals_b = [p2["prices"][m] for p2 in products_b if p2.get("prices", {}).get(m) is not None]
                tb = round(sum(vals_b), 2) if vals_b else None
                c5 = ws5.cell(row=row5, column=ci)
                c5.fill = PatternFill("solid", fgColor="1F3864")
                c5.alignment = center
                if ta and tb and tb != 0:
                    pct = (ta - tb) / tb
                    label = f"${ta:.2f}  /  ${tb:.2f}  ({'+' if pct>0 else ''}{pct*100:.1f}%)"
                    c5.value = label
                    c5.font  = Font(color="FF9999" if pct > 0 else "99FFB3", bold=True, size=10)
                else:
                    c5.value = f"${ta:.2f}" if ta else ""
                    c5.font  = Font(color="FFFFFF", bold=True, size=10)
            ws5.freeze_panes = "D3"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        fname = f"canasta_basica_{year}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Escaneo de facturas con IA ────────────────────────────────────────────────

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _match_catalog(detected_name: str, catalog: list) -> dict | None:
    best_score = 0.0
    best = None
    for p in catalog:
        score = _similarity(detected_name, p["name"])
        # También comparar contra palabras clave (p.ej. "BISTEC" debería matchear "RES BISTEC ESPECIAL")
        words = detected_name.lower().split()
        for w in words:
            if len(w) > 3:
                partial = _similarity(w, p["name"])
                score = max(score, partial * 0.85)
        if score > best_score:
            best_score = score
            best = p
    if best is None:
        return None
    confidence = "high" if best_score >= 0.72 else "medium" if best_score >= 0.45 else "low"
    return {**best, "score": round(best_score, 3), "confidence": confidence}


@router.post(
    "/api/canasta/{year}/scan-invoice",
    tags=["Canasta Básica"],
    summary="Escanear factura con IA (Gemini 3.5 Flash)",
    description=(
        "Envía una imagen de factura o ticket de supermercado al modelo gemini-3.5-flash. "
        "Detecta productos, precios y unidades. Hace fuzzy matching contra el catálogo del año. "
        "Retorna cada item con su nivel de confianza: "
        "high (score ≥ 0.72) · medium (≥ 0.45) · low (< 0.45). "
        "Autenticación: service_account.json vía google-genai enterprise=True — sin API key. "
        "Formatos soportados: JPEG · PNG · WEBP · HEIC."
    ),
)
async def scan_invoice(year: str, imagen: UploadFile = File(...)):
    if fdb is None:
        raise HTTPException(503, "Firestore no disponible")

    try:
        from google import genai as gai
        from google.genai import types as gtypes
    except ImportError:
        raise HTTPException(503, "Instala google-genai: pip install google-genai")

    try:
        import os
        from google.oauth2 import service_account as sa_mod

        image_bytes = await imagen.read()
        mime = imagen.content_type or "image/jpeg"

        sa_file = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "service_account.json")
        _creds = sa_mod.Credentials.from_service_account_file(
            sa_file,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
        client = gai.Client(
            enterprise=True,
            project="canaco-info",
            location="us",
            credentials=_creds,
        )

        prompt = (
            "Analiza esta imagen de una factura, ticket o recibo de supermercado mexicano. "
            "Extrae cada producto alimenticio o de higiene con su precio unitario y unidad de medida. "
            "Responde ÚNICAMENTE con un array JSON válido, sin texto adicional ni markdown. "
            "Ejemplo: [{\"nombre\": \"AGUACATE\", \"precio\": 39.95, \"unidad\": \"KILO\"}]"
        )

        response = client.models.generate_content(
            model="gemini-3.5-flash",
            contents=[
                gtypes.Part.from_bytes(data=image_bytes, mime_type=mime),
                prompt,
            ],
        )

        raw = response.text.strip()
        # Extraer el JSON aunque venga con ```json ... ```
        match = re.search(r'\[.*\]', raw, re.DOTALL)
        if not match:
            raise ValueError(f"Respuesta inesperada de Gemini: {raw[:200]}")

        detected = json.loads(match.group())
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al procesar imagen: {e}")

    # Cargar catálogo para hacer el matching
    catalog = _get_products(year)

    results = []
    for item in detected:
        nombre = str(item.get("nombre", "")).strip().upper()
        precio = item.get("precio")
        unidad = str(item.get("unidad", "")).strip().upper()
        if not nombre or precio is None:
            continue
        try:
            precio = float(precio)
        except (TypeError, ValueError):
            continue

        matched = _match_catalog(nombre, catalog)
        results.append({
            "detected_name": nombre,
            "detected_price": precio,
            "detected_unit": unidad,
            "matched_id":    matched["id"]         if matched else None,
            "matched_name":  matched["name"]       if matched else None,
            "confidence":    matched["confidence"] if matched else "low",
            "score":         matched["score"]      if matched else 0,
        })

    return {"ok": True, "items": results, "total_detected": len(results)}
